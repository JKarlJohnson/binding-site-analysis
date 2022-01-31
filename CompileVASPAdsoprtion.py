#COMPILE VASP ADSOPRTION DATA
#INPUT: VASP CONTCAR DATA, PARAMETER .txt FILE
#EVAN KERIN - EDIT 1.31.2022
#PYTHON 3.9.7

from ase import atoms
from ase.io import read, write, iread
import os, os.path
from openpyxl import Workbook
import math

#TOTAL FILES IN DIR
dirNum = (len([name for name in os.listdir('.') if os.path.isfile(name)]))
#ARRAY LIST OF FILES
dirList = os.listdir()


################  INITIALIZE SETTINGS FROM FILE ####################

ads_ATOM = []
surf_ATOM = []
OSZICAR = False

for item in range(dirNum):
    filename = str(dirList[item])
    if 'COMPILEPARAM' in filename:
        with open(filename) as f:
            lines = f.readlines()

#INITIALIZE PRESETS 
atopmin_Dist = 2.4
atopmax_Dist = 2.8
bridgedmin_Dist = 3.2
bridgedmax_Dist = 3.6
max_Dist = 6

#SETTING VARIABLES BASED ON KEYWORDS

for x in range(len(lines)):

    if 'ATOM_' in str(lines[x]):

        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        ads_ATOM.append(dataLineSplit[1])
        surf_ATOM.append(dataLineSplit[2])
    
    elif 'MAXDIST_' in str(lines[x]):
        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        max_Dist = float(dataLineSplit[1])

    elif 'ATOPMIN_' in str(lines[x]):
        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        atopmin_Dist = float(dataLineSplit[1])

    elif 'ATOPMAX_' in str(lines[x]):
        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        atopmax_Dist = float(dataLineSplit[1])

    elif 'BRIDGEDMIN_' in str(lines[x]):
        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        bridgedmin_Dist = float(dataLineSplit[1])

    elif 'BRIDGEDMAX_' in str(lines[x]):
        dataLine = str(lines[x]).strip()
        dataLineSplit = dataLine.split("_")

        bridgedmax_Dist = float(dataLineSplit[1])


    elif 'OSZICAR=TRUE' in str(lines[x]):
        OSZICAR = True



f.close

###################################################################

#DISTANCE CALC

def compute_Dist(xCord1,xCord2,yCord1,yCord2,zCord1,zCord2):

    atom_Dist = math.sqrt(math.pow(xCord1 - xCord2,2) + math.pow(yCord1 - yCord2,2) + math.pow(zCord1 - zCord2,2))

    return(atom_Dist)

#PULLING DATA FROM CONTCAR USING ASE THEN CHARACTERIZING BASED OFF DIST FUNC

def get_Contcar_Data(filename):

    dataList = []
    ads_FOUND = []
    surf_FOUND = []

    contcarFile = read(filename = filename)

    for x in range(len(contcarFile)):
        dataLine = str(contcarFile[x])
        dataLineSplit = dataLine.split(",")
        atomName = dataLineSplit[0].replace("Atom('","").replace("'","")

        if atomName in ads_ATOM:
            ads_FOUND.append(x)

        if atomName in surf_ATOM:
            surf_FOUND.append(x)

    
    for k in range(len(ads_FOUND)):
        
        dataLine1 = str(contcarFile[int(ads_FOUND[k])])
        dataLineSplit1 = dataLine1.split(",")
        atomName1 = str(k+1) + ":" + dataLineSplit1[0].replace("Atom('","").replace("'","")          
        xCord1 = float(dataLineSplit1[1].replace(" [",""))
        yCord1 = float(dataLineSplit1[2].replace(" ",""))
        zCord1 = float(dataLineSplit1[3].replace("]","").replace(" ",""))

        for j in range(len(surf_FOUND)):

            dataLine2 = str(contcarFile[int(surf_FOUND[j])])
            dataLineSplit2 = dataLine2.split(",")

            atomName2 = dataLineSplit2[0].replace("Atom('","").replace("'","")           
            xCord2 = float(dataLineSplit2[1].replace(" [",""))
            yCord2 = float(dataLineSplit2[2].replace(" ",""))
            zCord2 = float(dataLineSplit2[3].replace("]","").replace(" ",""))

            atom_Dist = compute_Dist(xCord1,xCord2,yCord1,yCord2,zCord1,zCord2)



            if atom_Dist < max_Dist:
                
                if atopmin_Dist and atopmax_Dist and bridgedmin_Dist and bridgedmax_Dist != None:

                    if atom_Dist > atopmin_Dist and atom_Dist < atopmax_Dist:
                        classifier = 'Atop'
                    
                    elif atom_Dist > bridgedmin_Dist and atom_Dist < bridgedmax_Dist:
                        classifier = 'Bridged'

                    elif atom_Dist > bridgedmax_Dist:
                        classifier = ''

                    elif atom_Dist < atopmin_Dist:
                        classifier = ''

                    else:
                        classifier = ''
            
                dataOutput = [filename, atomName1, atomName2, atom_Dist, classifier]
                dataList.append(dataOutput)
                dataList.sort(key=lambda x: x[3])

    return(dataList)



def getOSZICARE(filename):

    with open(filename, 'rb') as f:
        f.seek(-2, os.SEEK_END)
        while f.read(1) != b'\n':
            f.seek(-2, os.SEEK_CUR)
        energy_Data = f.readline().decode()
    
    return(energy_Data)




##################   MAIN    ####################

#CREATING XL WORKBOOK

wb = Workbook()
ws = wb.create_sheet('Data_Overview')
ws['A1'].value = "FileName"
ws['B1'].value = "Identifier & Adsorbate Atom"
ws['C1'].value = "Adsorbent Atom"
ws['D1'].value = "Distance (Angstrom)"
ws['E1'].value = "Classification"
ws['F1'].value = "Energies"

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 50

#CALLING FUNCTIONS 
xlrow = 2

for item in range(dirNum):
        filename = str(dirList[item])



        if '.CONTCAR' in filename:

            if OSZICAR == True:
                
                filenameFront = filename.replace(".CONTCAR","")
                
                for finder in range(dirNum):
                    filenameFinder = str(dirList[finder])

                    if filenameFront in filenameFinder:

                        OZfilename = filenameFront + ".OSZICAR"
                        energy_Data = getOSZICARE(OZfilename)
                        xlPlaceF = 'F' + str(xlrow)
                        ws[xlPlaceF].value = energy_Data




            dataList = get_Contcar_Data(filename)
           
            for i in range(len(dataList)):
                
                xlPlaceA = 'A' + str(xlrow)
                xlPlaceB = 'B' + str(xlrow)
                xlPlaceC = 'C' + str(xlrow)
                xlPlaceD = 'D' + str(xlrow)
                xlPlaceE = 'E' + str(xlrow)

                dataListLine = dataList[i]
               
                ws[xlPlaceA].value = dataListLine[0]
                ws[xlPlaceB].value = dataListLine[1]
                ws[xlPlaceC].value = dataListLine[2]
                ws[xlPlaceD].value = dataListLine[3]
                ws[xlPlaceE].value = dataListLine[4]

                xlrow = xlrow + 1



wb.save(filename = 'DataResults.xlsx')





